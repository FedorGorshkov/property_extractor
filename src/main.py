# External modules
import openpyxl as xl
from openpyxl.utils import range_boundaries, get_column_letter

# Project modules
from src import global_variables as gv
from src import methods
from src import funcs

# Built-in modules
import re

def get_possible_values(cell_address, xls_file, sheet):
    validator_name = None
    for validation in sheet.data_validations.dataValidation:
        ranges = validation.sqref.ranges
        for rng in ranges:
            if cell_address in rng:
                validator_name = validation.formula1
                break
    if validator_name is None:
        return []
    sheet_title, coord = [[sh, cd] for sh, cd in xls_file.defined_names[validator_name].destinations][0]
    res = []
    for rw in xls_file[sheet_title][coord]:
        for cll in rw:
            res.append(cll.value)
    return res


def main():
    # if len(sys.argv) < 2:
    #     print("Usage: python main.py xls_file")
    # xls_file = sys.argv[1]
    # try:
    #     xls_file = xl.load_workbook(xls_file)
    # except Exception as e:
    #     print(f"Failed to open {xls_file}")
    #     exit(-1)

    # DELETE: TEMPORARY FOR DEVELOPMENT
    # xls_file = "../xls/Tag_Property_Register_Template.xlsx"
    xls_file = "../intergaz/задание.xlsx"

    try:
        xls_file = xl.load_workbook(xls_file)
    except Exception as e:
        print(f"Failed to open {xls_file} ({e})")
        exit(-1)
    gv.init_spacy()
    gv.xls_file = xls_file
    sheet = xls_file[xls_file.sheetnames[0]]
    table = sheet.tables[list(sheet.tables.keys())[0]]
    columns = table.column_names
    min_col, min_row, max_col, max_row = range_boundaries(table.ref)
    if True in [name not in columns for name in gv.TARGET_COLUMNS]:
        print("\033[33m{}\033[0m".format(f'Some of columns {", ".join(gv.TARGET_COLUMNS)} not found in sheet'))
        exit(-1)
    Tag_index, description_index = columns.index(gv.TAG_COL), columns.index(gv.DESC_RU_COL)
    en_name_index, ru_name_index = columns.index(gv.NAME_COL), columns.index(gv.NAME_RU_COL)
    value_index, UM_index = columns.index(gv.VALUE_COL), columns.index(gv.UM_COL)
    to_find, to_skip = dict(), list()
    for row in range(min_row + 1, max_row + 1):
        tag = sheet[get_column_letter(Tag_index + 1) + str(row)].value
        if not tag:
            continue
        if tag.startswith("E-"):
            tag = tag[2:]
        if tag in to_skip:
            continue
        ids = gv.cur.execute("SELECT file_id FROM files WHERE mentioned_tags LIKE ?", (f"%{tag}%", )).fetchall()
        if not ids:
            to_skip.append(tag)
            if gv.LOGS:
                print(f"Tag \"{tag}\" not found in file bank, skipping")
            continue
        desc = sheet[get_column_letter(description_index + 1) + str(row)].value
        property_name_ru = sheet[get_column_letter(ru_name_index + 1) + str(row)].value
        property_name_en = sheet[get_column_letter(en_name_index + 1) + str(row)].value
        to_add = {"en_name": re.sub(r'\([^()]*\)', '', property_name_en.replace('.', '-')),
                  "ru_name": re.sub(r'\([^()]*\)', '', property_name_ru.replace('.', '-')),
                  "desc": desc, "idx": str(row), "files": set(ids[0])}
        if tag not in to_find:
            to_find[tag] = [to_add]
        else:
            to_find[tag].append(to_add)

    # new_to_find = dict()
    # include_only = {
    #     "7350-SJ-0001B": ["pressure rating", "driver type", "lower limit operating inlet temperature", "nominal inlet diameter",
    #                       "po issuer company name", "fluid name", "upper limit operating inlet temperature", "po code",
    #                       "design lifetime", "fluid phase", ],
    #     "8500-SS-01-T-SX-2001": ["telecom equipment installation type", "po issuer company name", "cabinet mounting method",
    #                              "height in unit", "po code", "design lifetime", ],
    #     "72-P-4001A": ["driver equipment", "upper limit operating volume flow rate", "driver type", "fluid name"
    #                    "normal operating outlet pressure", "normal operating inlet pressure", "normal operating liquid density",
    #                    "normal operating rotational speed", "normal operating dynamic viscosity", "upper limit design pressure",
    #                    "normal operating temperature", "po issuer company name", "po code", "design lifetime",
    #                    "explosion protection gas group required", "upper limit allowable sound pressure level"]
    # }
    # to_find = {"72-P-4001A": to_find["72-P-4001A"]}
    # del to_find["72-P-4001A"]
    # del to_find["7350-SJ-0001B"]
    # for tag in include_only:
    #     new_to_find[tag] = []
    #     for prop in to_find[tag]:
    #         if prop["en_name"] in include_only[tag]:
    #             new_to_find[tag].append(prop)
    # for tag in to_find:
    #     new_to_find[tag] = [to_find[tag][0]]
    # to_find = new_to_find
    if gv.LOGS:
        with open("logs.txt", "w") as f:
            f.write(f"Actual objects to find: {list(to_find.keys())}\n")
        gv.original_print()
        print(f"Actual objects to find: {list(to_find.keys())}")
        print(f"Amount tasks: {sum([len(to_find[tag]) for tag in to_find])}")
        gv.original_print()

    gv.ALL_TAGS = gv.ALL_TAGS.union(list(set(to_find.keys())))
    failed, success = [], 0
    # MAIN CYCLE ГЛАВНЫЙ ЦИКЛ
    for tag in to_find:
        if not to_find[tag]:
            continue
        if gv.LOGS:
            print(f"Now executing tasks for tag {tag}")
            with open("logs.txt", "a") as f:
                f.write(f"Now executing tasks for tag {tag}\n")
        files_set = set()
        for task in to_find[tag]:
            files_set = files_set.union(set(task["files"]))
        files, overall_len = dict(), [0, 0, 0, 0]
        for file in files_set:
            files[file] = funcs.preload_file(file, tag)
            overall_len = [ overall_len[0] + len( files[file]["pure_text"] ),
                            overall_len[1] + len( files[file]["tables"] ),
                            overall_len[2] + ( len( files[file]["tech_specs"] ) if isinstance(files[file]["tech_specs"], str) else 0 ),
                            overall_len[3] + ( 1 if files[file]["passport"] else 0 )    ]
        if gv.LOGS:
            print(f"Pre-loaded file(-s) with numbers: {files_set}; Relevant length of text {overall_len[0]}; Relevant tables {overall_len[1]}; Tech specs with len {overall_len[2]}; Passports {overall_len[3]};")
            with open("logs.txt", "a") as f:
                f.write(f"Pre-loaded file(-s) with numbers: {files_set}; Relevant length of text {overall_len[0]}; Relevant tables {overall_len[1]}; Tech specs with len {overall_len[2]}; Passports {overall_len[3]};\n")
        for task in to_find[tag]:
            if gv.LOGS:
                print(f"\tNow searching for \"{task['en_name']}\" in {len(task['files'])} file(-s)")
                with open("logs.txt", "a") as f:
                    f.write(f"\tNow searching for \"{task['en_name']}\"\n")
            possible_values = get_possible_values(get_column_letter(value_index + 1) + task["idx"], xls_file, sheet)
            possible_uoms = get_possible_values(get_column_letter(UM_index + 1) + task["idx"], xls_file, sheet)
            res = gv.cur.execute("SELECT prop_desc_eng, condition, alt_keywords, desc_for_llm, flag FROM properties WHERE prop_name_eng = ?", (task["en_name"],)).fetchone()
            condition, alt_keywords, desc_for_llm, flag = "none", "none", "", 0
            if res:
                if len(res) > 2:
                    some_desc, condition, alt_keywords, desc_for_llm, flag = res[0], res[1], res[2], res[3], res[4]
                    if not task["desc"] and some_desc != "none":
                        task["desc"] = some_desc
            if gv.LOGS:
                if possible_values:
                    print(f"\t\tPossible values: {possible_values}")
                if possible_uoms:
                    print(f"\t\tPossible UoMs: {possible_uoms}")
                if condition not in ["none", None]:
                    print(f"\t\tCondition: {condition}")
            original_prop_name = task["en_name"]
            solver = methods.Solver(tag, task, possible_values, possible_uoms, condition=condition,
                                    alt_keywords=alt_keywords.split(";")[:-1] if alt_keywords != "none" else None,
                                    desc_for_llm=desc_for_llm, flag=flag,)
            for relevant_file in task["files"]:
                solver.set_file(files[relevant_file])
                solver.solve()
            if not gv.validate_result(solver.result):
                if gv.LOGS:
                    print("\033[31m{}\033[0m".format(f"\t\tCouldn't find prop '{original_prop_name}' in file bank"))
                failed.append(task)
            else:
                gv.MODIFIED = True
                cell = sheet[get_column_letter(value_index + 1) + task["idx"]]
                cell.value = solver.result
                if isinstance(solver.uom, str):
                    if not (solver.uom.startswith(gv.LLM_FAILURE_FLAG) or solver.uom.endswith(gv.LLM_FAILURE_FLAG)):
                        cell = sheet[get_column_letter(UM_index + 1) + task["idx"]]
                        cell.value = solver.uom
                success += 1
                print("\033[32m{}\033[0m".format(f"\t\tSuccess! Found value {solver.result} and UoM {solver.uom}"))
            if gv.LOGS:
                print('   ', '-' * 80)

    print(f"Success: {success}, failed: {len(failed)}")
    print(f"Time spent for requests: {gv.REQUESTS_TIME_SPENT // 60} minutes and {round(gv.REQUESTS_TIME_SPENT % 60, 2)} seconds")
    print(f"Time spent for Solver.methods: {gv.METHODS_TIME_SPENT // 60} minutes and {round(gv.METHODS_TIME_SPENT % 60, 2)} seconds")
    xls_file.save(f"../xls/Test_output.xlsx")

if __name__ == "__main__":
    main()
